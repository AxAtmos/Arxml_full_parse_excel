import xml.etree.ElementTree as ET
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import os

NS = {'ns': 'http://autosar.org/schema/r4.0'}

def parse_arxml_element(element, current_path_short_names, all_parsed_data, max_depth):
    row_data = {}
    tag = element.tag.split('}')[-1] if '}' in element.tag else element.tag
    row_data['TAG'] = tag
    
    short_name_element = element.find('ns:SHORT-NAME', NS)
    short_name = short_name_element.text if short_name_element is not None else ''
    row_data['SHORT-NAME'] = short_name
    
    dest = ''
    text_content = ''

    if 'DEST' in element.attrib:
        dest = element.attrib['DEST']
    
    if element.text and element.text.strip():
        text_content = element.text.strip()

    row_data['DEST'] = dest
    row_data['TEXT'] = text_content

    current_element_path = current_path_short_names + [short_name]

    if len(current_element_path) > max_depth[0]:
        max_depth[0] = len(current_element_path)

    for i, sn in enumerate(current_element_path):
        row_data[f'Level {i+1} Short-Name'] = sn

    all_parsed_data.append(row_data)

    for child in element:
        child_tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if child_tag in ['SHORT-NAME']:
            continue
        parse_arxml_element(child, current_element_path, all_parsed_data, max_depth)

def apply_excel_styles(sheet):

    sheet.freeze_panes = 'A2'
    sheet.row_dimensions[1].height = 30

    header_font = Font(name='Arial', size=11, color='4D4D4D', bold=True)
    header_fill = PatternFill(start_color='63EFCE', end_color="63EFCE", fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    content_font = Font(name='Arial', size=11, color="202020")
    content_alignment = Alignment(vertical='center')

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.font = content_font
            cell.alignment = content_alignment

    for col in range(1, sheet.max_column + 1):
        col_letter = get_column_letter(col)
        sheet.column_dimensions[col_letter].alignment = content_alignment

def parse_arxml_to_excel(arxml_file_path, output_excel_path):
    try:
        tree = ET.parse(arxml_file_path)
        root = tree.getroot()
    except ET.ParseError as e:
        print(f"错误：解析ARXML文件失败: {e}")
        return
    except FileNotFoundError:
        print(f"错误：文件未找到: {arxml_file_path}")
        return

    workbook = openpyxl.Workbook()
    print(f"OpenPyxl内容已构建，开始执行解析。")

    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])

    top_level_ar_packages = []
    ar_packages_container = root.find('ns:AR-PACKAGES', NS)
    if ar_packages_container is not None:
        top_level_ar_packages = ar_packages_container.findall('ns:AR-PACKAGE', NS)

    if not top_level_ar_packages:
        print("未在ARXML文件中找到顶层AR-PACKAGE元素。尝试将整个文档解析到单个表中。")
        sheet_name = "ARXML_Data"
        all_parsed_data_for_sheet = []
        max_depth_for_sheet = [0]

        parse_arxml_element(root, [], all_parsed_data_for_sheet, max_depth_for_sheet)
        
        if not all_parsed_data_for_sheet:
            print("未从整个文档中解析到任何数据。")
            return

        sheet = workbook.create_sheet(title=sheet_name)

        headers = [f'Level {i+1} Short-Name' for i in range(max_depth_for_sheet[0])]
        headers.extend(['TAG', 'SHORT-NAME', 'DEST', 'TEXT'])
        sheet.append(headers)

        for row_dict in all_parsed_data_for_sheet:
            row_values = []
            for header in headers:
                row_values.append(row_dict.get(header, ''))
            sheet.append(row_values)

        for col_idx, header in enumerate(headers, 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in sheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        apply_excel_styles(sheet)
        workbook.save(output_excel_path)
        print(f"ARXML文件已成功解析并保存到 {output_excel_path}")
        return

    for ar_package in top_level_ar_packages:
        package_short_name_element = ar_package.find('ns:SHORT-NAME', NS)
        if package_short_name_element is None:
            print(f"跳过一个没有SHORT-NAME的AR-PACKAGE。")
            continue
        
        sheet_name = package_short_name_element.text
        sheet_name = sheet_name[:31].replace('/', '_').replace('\\', '_').replace('*', '_').replace('?', '_').replace('[', '_').replace(']', '_').replace(':', '_')
        
        original_sheet_name = sheet_name
        counter = 1
        while sheet_name in workbook.sheetnames:
            sheet_name = f"{original_sheet_name}_{counter}"
            counter += 1

        all_parsed_data_for_sheet = []
        max_depth_for_sheet = [0]

        parse_arxml_element(ar_package, [], all_parsed_data_for_sheet, max_depth_for_sheet)

        if not all_parsed_data_for_sheet:
            print(f"表 '{sheet_name}' 没有解析到任何数据。")
            continue

        sheet = workbook.create_sheet(title=sheet_name)

        headers = [f'Level {i+1} Short-Name' for i in range(max_depth_for_sheet[0])]
        headers.extend(['TAG', 'SHORT-NAME', 'DEST', 'TEXT'])
        sheet.append(headers)

        for row_dict in all_parsed_data_for_sheet:
            row_values = []
            for header in headers:
                row_values.append(row_dict.get(header, ''))
            sheet.append(row_values)
        
        for col_idx, header in enumerate(headers, 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for cell in sheet[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except TypeError:
                    pass
            adjusted_width = (max_length + 5)
            sheet.column_dimensions[column].width = adjusted_width

        apply_excel_styles(sheet)
        print(f"表 {sheet_name} 已构建完毕！")
        print(f"格式已渲染，共 {len(all_parsed_data_for_sheet)} 行 {col_idx} 列数据。")

    workbook.save(output_excel_path)
    print(f"ARXML文件已成功解析并保存到 {output_excel_path}")

if __name__ == "__main__":
    arxml_input_file = input("请输入ARXML文件的完整路径 (例如: C:\\Users\\XXX\\Desktop\\arxml_file.arxml): ")
    excel_output_file = input("请输入输出Excel文件的名称 (例如: output.xlsx): ")

    if not os.path.isabs(excel_output_file):
        output_directory = os.path.dirname(arxml_input_file)
        excel_output_file = os.path.join(output_directory, excel_output_file)

    parse_arxml_to_excel(arxml_input_file, excel_output_file)